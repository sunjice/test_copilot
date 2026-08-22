"""用例域 — 业务逻辑层：项目/套件/用例 CRUD + Excel 导入 + 审核工作台 + AI 结果写入。"""

import json
import re
from datetime import datetime
from io import BytesIO
from typing import Any

import openpyxl
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession
from loguru import logger

from app.exceptions import BusinessException
from app.pagination import PageResult
from app.response import ResultCode
from app.aitc.case.models import AiTcProject, AiTcSuite, AiTcCase
from app.aitc.task.models import AiTcTaskItem, AiTcReviewRecord
from app.aitc.case.schemas import (
    CaseCoreMark, CaseQuery, CaseStep, CaseUpdate, CaseVO,
    CaseSampleMark,
    CaseReviewReq, CaseReviewDetailVO, CaseFieldReviewItem,
    FieldSuggestionVO, PendingCaseVO, PendingSuiteNodeVO,
    ImportResult,
    ProjectCreate, ProjectQuery, ProjectUpdate, ProjectVO,
    SuiteNodeVO, SuiteVO,
)
from app.aitc.schemas import OptionVO
from app.aitc.constants import (
    CaseImportance, CoreSource, ReviewStatus,
    ConfirmStatus, ItemStatus,
)


class CaseService:
    """用例域全部业务逻辑（项目/套件/用例/审核/AI结果写入）。"""

    def __init__(self, db: AsyncSession):
        self.db = db

    # ═══════════════ 项目 CRUD ═══════════════

    async def get_project_page(self, query: ProjectQuery) -> PageResult:
        conditions = [AiTcProject.is_deleted == 0]
        if query.keywords:
            kw = f"%{query.keywords}%"
            conditions.append(AiTcProject.name.ilike(kw) | AiTcProject.prefix.ilike(kw))

        stmt = select(AiTcProject).where(*conditions).order_by(AiTcProject.id.desc())
        count_q = select(func.count()).select_from(stmt.subquery())
        total = (await self.db.execute(count_q)).scalar() or 0

        offset = (query.pageNum - 1) * query.pageSize
        rows = await self.db.execute(stmt.offset(offset).limit(query.pageSize))
        items = rows.scalars().all()

        return PageResult(
            records=[self._project_to_vo(p) for p in items],
            total=total, pageNum=query.pageNum, pageSize=query.pageSize,
        )

    async def get_project_options(self) -> list[OptionVO]:
        rows = await self.db.execute(
            select(AiTcProject.id, AiTcProject.name)
            .where(AiTcProject.is_deleted == 0)
            .order_by(AiTcProject.id)
        )
        return [OptionVO(value=r.id, label=r.name) for r in rows]

    async def get_project_by_id(self, pid: int) -> ProjectVO:
        result = await self.db.execute(
            select(AiTcProject).where(AiTcProject.id == pid, AiTcProject.is_deleted == 0)
        )
        p = result.scalar_one_or_none()
        if p is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="项目不存在")
        return self._project_to_vo(p)

    async def create_project(self, form: ProjectCreate) -> ProjectVO:
        exist = await self.db.execute(
            select(AiTcProject).where(AiTcProject.prefix == form.prefix)
        )
        existing = exist.scalar_one_or_none()
        if existing is not None:
            if existing.is_deleted == 0:
                raise BusinessException(code=ResultCode.DUPLICATE_KEY, msg="项目标识已存在")
            existing.name = form.name
            existing.prefix = form.prefix
            existing.description = form.description
            existing.is_deleted = 0
            existing.update_time = datetime.now()
            await self.db.flush()
            logger.info(f"AiTcProject restored: {form.name}")
            return self._project_to_vo(existing)

        p = AiTcProject(**form.model_dump())
        self.db.add(p)
        await self.db.flush()
        logger.info(f"AiTcProject created: {form.name} id={p.id}")
        return self._project_to_vo(p)

    async def update_project(self, form: ProjectUpdate) -> ProjectVO:
        result = await self.db.execute(
            select(AiTcProject).where(AiTcProject.id == form.id, AiTcProject.is_deleted == 0)
        )
        p = result.scalar_one_or_none()
        if p is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="项目不存在")
        exist = await self.db.execute(
            select(AiTcProject.id).where(
                AiTcProject.prefix == form.prefix,
                AiTcProject.id != form.id,
                AiTcProject.is_deleted == 0,
            )
        )
        if exist.scalar() is not None:
            raise BusinessException(code=ResultCode.DUPLICATE_KEY, msg="项目标识已存在")
        p.name = form.name
        p.prefix = form.prefix
        p.description = form.description
        p.update_time = datetime.now()
        await self.db.flush()
        return self._project_to_vo(p)

    async def delete_project(self, ids: str) -> int:
        id_list = [int(x) for x in ids.split(",") if x.strip()]
        if not id_list:
            raise BusinessException(code=ResultCode.PARAM_VALID_FAIL, msg="请选择项目")
        await self.db.execute(
            text("UPDATE ai_tc_projects SET is_deleted = 1 WHERE id = ANY(:ids)"),
            {"ids": id_list},
        )
        return len(id_list)

    # ═══════════════ 套件树 ═══════════════

    async def get_suite_tree(self, project_id: int) -> list[SuiteNodeVO]:
        """获取项目下套件树（仅套件层级，不含用例节点）。"""
        rows = await self.db.execute(
            select(AiTcSuite)
            .where(AiTcSuite.project_id == project_id, AiTcSuite.is_deleted == 0)
            .order_by(AiTcSuite.sort_order, AiTcSuite.id)
        )
        suites = rows.scalars().all()

        # 批量查各节点用例数
        suite_ids = [s.id for s in suites]
        case_counts: dict[int, int] = {}
        if suite_ids:
            cnt_rows = await self.db.execute(
                select(AiTcCase.suite_id, func.count(AiTcCase.id))
                .where(AiTcCase.suite_id.in_(suite_ids), AiTcCase.is_deleted == 0)
                .group_by(AiTcCase.suite_id)
            )
            case_counts = {r.suite_id: r[1] for r in cnt_rows}

        vo_map: dict[int, SuiteNodeVO] = {}
        for s in suites:
            vo = SuiteNodeVO(
                id=s.id, label=s.name, name=s.name,
                description=s.description,
                project_id=s.project_id, parent_id=s.parent_id,
                sort_order=s.sort_order, case_count=case_counts.get(s.id, 0),
            )
            vo_map[s.id] = vo

        tree = []
        for vo in vo_map.values():
            parent = vo_map.get(vo.parent_id)
            if parent:
                parent.children.append(vo)
            else:
                tree.append(vo)

        # 递归累加子节点用例数
        def _accumulate(node: SuiteNodeVO) -> int:
            for child in node.children:
                node.case_count += _accumulate(child)
            return node.case_count
        for root in tree:
            _accumulate(root)

        return tree

    async def get_suite_children(
        self, suite_id: int, project_id: int | None = None,
    ) -> list[SuiteNodeVO]:
        """获取套件子节点（子套件 + 直接用例），供树懒加载。
        suite_id=0 时返回项目根级套件。
        """
        children: list[SuiteNodeVO] = []

        # 取项目前缀
        project_prefix = ""
        if suite_id > 0:
            suite = await self.db.get(AiTcSuite, suite_id)
            if suite and suite.project_id:
                proj = await self.db.get(AiTcProject, suite.project_id)
                if proj:
                    project_prefix = proj.prefix
        elif project_id is not None:
            proj = await self.db.get(AiTcProject, project_id)
            if proj:
                project_prefix = proj.prefix

        # 1. 子套件
        child_filter = [AiTcSuite.parent_id == suite_id, AiTcSuite.is_deleted == 0]
        if suite_id == 0 and project_id is not None:
            child_filter.append(AiTcSuite.project_id == project_id)
        child_rows = await self.db.execute(
            select(AiTcSuite)
            .where(*child_filter)
            .order_by(AiTcSuite.sort_order, AiTcSuite.id)
        )
        child_suites = child_rows.scalars().all()

        # 递归统计子套件用例数（含后代子模块）
        child_suite_ids = [s.id for s in child_suites]
        case_counts: dict[int, int] = {}
        if child_suite_ids:
            suite_subtree_ids: dict[int, list[int]] = {}
            all_ids: list[int] = []
            for s in child_suites:
                subtree = await self._get_subtree_suite_ids(s.id)
                suite_subtree_ids[s.id] = subtree
                all_ids.extend(subtree)
            if all_ids:
                cnt_rows = await self.db.execute(
                    select(AiTcCase.suite_id, func.count(AiTcCase.id))
                    .where(AiTcCase.suite_id.in_(all_ids), AiTcCase.is_deleted == 0)
                    .group_by(AiTcCase.suite_id)
                )
                db_counts = {r.suite_id: r[1] for r in cnt_rows}
                for s in child_suites:
                    subtree = suite_subtree_ids.get(s.id, [s.id])
                    case_counts[s.id] = sum(db_counts.get(sid, 0) for sid in subtree)

        for s in child_suites:
            children.append(SuiteNodeVO(
                id=s.id, label=s.name, name=s.name,
                description=s.description,
                project_id=s.project_id, project_prefix=project_prefix,
                parent_id=s.parent_id,
                sort_order=s.sort_order, case_count=case_counts.get(s.id, 0),
            ))

        # 2. 直接用例（仅 suite_id > 0，限 200 条）
        if suite_id > 0:
            direct_cnt = await self.db.scalar(
                select(func.count(AiTcCase.id))
                .where(AiTcCase.suite_id == suite_id, AiTcCase.is_deleted == 0)
            ) or 0
            if direct_cnt > 0:
                case_rows = await self.db.execute(
                    select(AiTcCase.id, AiTcCase.external_id, AiTcCase.name, AiTcCase.purpose)
                    .where(AiTcCase.suite_id == suite_id, AiTcCase.is_deleted == 0)
                    .order_by(AiTcCase.id)
                    .limit(200)
                )
                suite = child_suites[0] if child_suites else None
                proj_id = suite.project_id if suite else None
                for cid, ext_id, cname, cpurpose in case_rows:
                    children.append(SuiteNodeVO(
                        id=-cid,
                        label=f"{project_prefix}{ext_id or ''}__{cname}",
                        name=cname,
                        project_id=proj_id,
                        project_prefix=project_prefix,
                        parent_id=suite_id,
                        node_type="case",
                        external_id=ext_id,
                        case_count=0,
                    ))

        return children

    async def _get_subtree_suite_ids(self, suite_id: int) -> list[int]:
        """获取指定套件及其所有子套件的 ID 列表（纯 parent_id 递归，不依赖 tree_path）。"""
        ids: list[int] = [suite_id]
        current: list[int] = [suite_id]
        while current:
            rows = await self.db.execute(
                select(AiTcSuite.id).where(
                    AiTcSuite.parent_id.in_(current),
                    AiTcSuite.is_deleted == 0,
                )
            )
            children = [r[0] for r in rows]
            if not children:
                break
            ids.extend(children)
            current = children
        return ids

    # ═══════════════ 用例 CRUD ═══════════════

    async def get_case_page(self, query: CaseQuery) -> PageResult:
        conditions = [AiTcCase.is_deleted == 0]
        if query.projectId is not None:
            conditions.append(AiTcCase.project_id == query.projectId)
        if query.suiteId is not None:
            conditions.append(AiTcCase.suite_id == query.suiteId)
        if query.isCore is not None:
            conditions.append(AiTcCase.is_core == query.isCore)
        if query.isSample is not None:
            conditions.append(AiTcCase.is_sample == query.isSample)
        if query.reviewStatus is not None:
            conditions.append(AiTcCase.review_status == query.reviewStatus)
        if query.importance is not None:
            conditions.append(AiTcCase.importance == query.importance)
        if query.keywords:
            kw = f"%{query.keywords}%"
            conditions.append(
                AiTcCase.name.ilike(kw) | AiTcCase.external_id.ilike(kw)
            )

        # 动态排序
        sort_cols = {
            "external_id": AiTcCase.external_id,
            "importance": AiTcCase.importance,
            "is_core": AiTcCase.is_core,
        }
        order_clauses = []
        if query.sortField and query.sortField in sort_cols:
            col = sort_cols[query.sortField]
            order_clauses.append(col.desc() if query.sortOrder == "descending" else col.asc())
        order_clauses.extend([AiTcCase.suite_id, AiTcCase.id])

        stmt = (
            select(AiTcCase)
            .where(*conditions)
            .order_by(*order_clauses)
        )
        count_q = select(func.count()).select_from(stmt.subquery())
        total = (await self.db.execute(count_q)).scalar() or 0

        offset = (query.pageNum - 1) * query.pageSize
        rows = await self.db.execute(stmt.offset(offset).limit(query.pageSize))
        cases = rows.scalars().all()

        # 批量查套件名 & 项目前缀
        suite_ids = list({c.suite_id for c in cases})
        suite_map: dict[int, str] = {}
        if suite_ids:
            s_rows = await self.db.execute(
                select(AiTcSuite.id, AiTcSuite.name).where(AiTcSuite.id.in_(suite_ids))
            )
            suite_map = {r.id: r.name for r in s_rows}

        proj_ids = list({c.project_id for c in cases})
        prefix_map: dict[int, str] = {}
        if proj_ids:
            p_rows = await self.db.execute(
                select(AiTcProject.id, AiTcProject.prefix).where(AiTcProject.id.in_(proj_ids))
            )
            prefix_map = {r.id: r.prefix for r in p_rows}

        return PageResult(
            records=[self._case_to_vo(c, suite_map.get(c.suite_id), prefix_map.get(c.project_id, "")) for c in cases],
            total=total, pageNum=query.pageNum, pageSize=query.pageSize,
        )

    async def get_case_by_id(self, case_id: int) -> CaseVO:
        result = await self.db.execute(
            select(AiTcCase).where(AiTcCase.id == case_id, AiTcCase.is_deleted == 0)
        )
        c = result.scalar_one_or_none()
        if c is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="用例不存在")
        suite_name = ""
        if c.suite_id:
            s = await self.db.get(AiTcSuite, c.suite_id)
            if s:
                suite_name = s.name
        project_prefix = ""
        if c.project_id:
            proj = await self.db.get(AiTcProject, c.project_id)
            if proj:
                project_prefix = proj.prefix
        return self._case_to_vo(c, suite_name, project_prefix)

    async def update_case(self, case_id: int, form: CaseUpdate) -> CaseVO:
        result = await self.db.execute(
            select(AiTcCase).where(AiTcCase.id == case_id, AiTcCase.is_deleted == 0)
        )
        c = result.scalar_one_or_none()
        if c is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="用例不存在")
        c.external_id = form.external_id
        c.name = form.name
        c.purpose = form.purpose
        c.summary = form.summary
        c.preconditions = form.preconditions
        c.topo = form.topo
        c.test_data = form.test_data
        c.steps = [s.model_dump() for s in form.steps]
        c.importance = form.importance
        c.update_time = datetime.now()
        await self.db.flush()
        suite_name = ""
        if c.suite_id:
            s = await self.db.get(AiTcSuite, c.suite_id)
            if s:
                suite_name = s.name
        project_prefix = ""
        if c.project_id:
            proj = await self.db.get(AiTcProject, c.project_id)
            if proj:
                project_prefix = proj.prefix
        return self._case_to_vo(c, suite_name, project_prefix)

    async def mark_case_core(self, form: CaseCoreMark) -> None:
        """人工标记/取消核心用例。"""
        result = await self.db.execute(
            select(AiTcCase).where(AiTcCase.id == form.case_id, AiTcCase.is_deleted == 0)
        )
        c = result.scalar_one_or_none()
        if c is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="用例不存在")
        c.is_core = form.is_core
        c.core_reason = form.reason if form.is_core else None
        c.core_source = CoreSource.MANUAL if form.is_core else None
        c.update_time = datetime.now()
        await self.db.flush()

    async def mark_case_sample(self, form: CaseSampleMark) -> None:
        """人工标记/取消样本用例。"""
        result = await self.db.execute(
            select(AiTcCase).where(AiTcCase.id == form.case_id, AiTcCase.is_deleted == 0)
        )
        c = result.scalar_one_or_none()
        if c is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="用例不存在")
        c.is_sample = form.is_sample
        c.update_time = datetime.now()
        await self.db.flush()

    async def delete_cases(self, ids: str) -> int:
        id_list = [int(x) for x in ids.split(",") if x.strip()]
        if not id_list:
            raise BusinessException(code=ResultCode.PARAM_VALID_FAIL, msg="请选择用例")
        await self.db.execute(
            text("UPDATE ai_tc_cases SET is_deleted = 1 WHERE id = ANY(:ids)"),
            {"ids": id_list},
        )
        return len(id_list)

    # ═══════════════ Excel 导入 ═══════════════

    async def import_cases(self, project_id: int, file_content: bytes) -> ImportResult:
        """解析 Excel 文件并导入用例。"""
        proj = await self.db.get(AiTcProject, project_id)
        if proj is None or proj.is_deleted == 1:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="项目不存在")

        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

        created, updated = 0, 0
        errors: list[dict] = []
        suite_cache: dict[str, int] = {}  # path -> suite_id，缓存避免重复查库

        for idx, row in enumerate(rows, start=2):
            try:
                if not row or all(v is None for v in row):
                    continue
                (external_id, module_path, name, level, summary,
                 preconditions, topo, test_data, steps_raw, expected_raw) = (
                    self._safe_str(row, i) for i in range(10)
                )

                if not external_id or not name:
                    errors.append({"row": idx, "msg": "用例ID或名称为空"})
                    continue

                suite_id = await self._ensure_suite(project_id, module_path, suite_cache)
                importance = CaseImportance.from_label(level) if level else 2
                step_list = self._parse_steps(steps_raw, expected_raw)

                case_data = dict(
                    project_id=project_id, suite_id=suite_id,
                    external_id=str(external_id).strip(), name=name.strip(),
                    summary=summary.strip() if summary else None,
                    preconditions=preconditions.strip() if preconditions else None,
                    topo=topo.strip() if topo else None,
                    test_data=test_data.strip() if test_data else None,
                    steps=step_list, importance=importance,
                )

                exist = await self.db.execute(
                    select(AiTcCase).where(
                        AiTcCase.project_id == project_id,
                        AiTcCase.external_id == case_data["external_id"],
                        AiTcCase.is_deleted == 0,
                    )
                )
                existing = exist.scalar_one_or_none()
                if existing:
                    for k, v in case_data.items():
                        setattr(existing, k, v)
                    existing.update_time = datetime.now()
                    updated += 1
                else:
                    self.db.add(AiTcCase(**case_data))
                    created += 1

            except Exception as e:
                errors.append({"row": idx, "msg": str(e)})

        await self.db.flush()

        # 更新项目最后同步时间
        proj.last_sync_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        proj.update_time = datetime.now()
        await self.db.flush()

        logger.info(f"Excel import done: created={created} updated={updated} errors={len(errors)}")
        return ImportResult(created=created, updated=updated, errors=errors)

    async def download_template(self) -> bytes:
        """生成导入模板 xlsx。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用例导入模板"
        headers = ["用例ID", "所属模块", "用例名称", "级别", "测试思想", "预置条件", "测试Topo", "测试数据", "测试步骤", "预期结果"]
        ws.append(headers)
        ws.append(["TC-001", "认证模块/登录", "正确账号密码登录", "高", "验证登录主流程",
                    "已注册账号，用户未登录", "PC直连DUT", "账号admin/密码123456",
                    "1. 打开登录页面\n2. 输入账号密码\n3. 点击登录按钮",
                    "1. 显示登录表单\n2. 输入框填充成功\n3. 跳转首页，显示用户名"])

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def _ensure_suite(self, project_id: int, module_path: str | None, cache: dict[str, int]) -> int:
        """按路径层级查找/创建套件，返回最终的叶子套件 ID。"""
        if not module_path:
            module_path = "未分类"
        parts = [p.strip() for p in module_path.split("/") if p.strip()]
        parent_id = 0
        current_path = ""
        for part in parts:
            current_path = f"{current_path}/{part}" if current_path else part
            if current_path in cache:
                parent_id = cache[current_path]
                continue
            exist = await self.db.execute(
                select(AiTcSuite).where(
                    AiTcSuite.project_id == project_id,
                    AiTcSuite.parent_id == parent_id,
                    AiTcSuite.name == part,
                    AiTcSuite.is_deleted == 0,
                )
            )
            suite = exist.scalar_one_or_none()
            if suite:
                parent_id = suite.id
            else:
                if parent_id == 0:
                    tree_path = "0"
                else:
                    parent_suite = await self.db.get(AiTcSuite, parent_id)
                    tree_path = f"{parent_suite.tree_path}{parent_id}," if parent_suite else "0"
                suite = AiTcSuite(
                    project_id=project_id, parent_id=parent_id,
                    tree_path=f"{tree_path}{parent_id}," if parent_id else tree_path,
                    name=part, sort_order=0,
                )
                self.db.add(suite)
                await self.db.flush()
                if parent_id == 0:
                    suite.tree_path = "0"
                else:
                    suite.tree_path = f"{tree_path}{parent_id},"
                await self.db.flush()
                parent_id = suite.id
            cache[current_path] = parent_id
        return parent_id

    def _parse_steps(self, steps_raw: str, expected_raw: str) -> list[dict]:
        """将单元格内的多行步骤文本解析为 [{step_no, action, expected}]。"""
        actions = self._split_lines(steps_raw or "")
        expected = self._split_lines(expected_raw or "")
        result = []
        max_len = max(len(actions), len(expected))
        for i in range(max_len):
            result.append({
                "step_no": i + 1,
                "action": actions[i] if i < len(actions) else "",
                "expected": expected[i] if i < len(expected) else "",
            })
        return result

    @staticmethod
    def _split_lines(text: str) -> list[str]:
        """按换行拆分并去除行首编号。"""
        lines = [l.strip() for l in re.split(r"[\r\n]+", text) if l.strip()]
        cleaned = []
        for line in lines:
            line = re.sub(r"^[\s]*[(（]?\d+[)）]?[\.、]?\s*", "", line)
            if line:
                cleaned.append(line)
        return cleaned

    @staticmethod
    def _safe_str(row, idx) -> str:
        if idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    # ═══════════════ 审核工作台 ═══════════════

    async def get_pending_review_tree(self, project_id: int) -> list[PendingSuiteNodeVO]:
        """获取项目下套件树，每个节点标注待审核的用例数。"""
        rows = await self.db.execute(
            select(AiTcSuite)
            .where(AiTcSuite.project_id == project_id, AiTcSuite.is_deleted == 0)
            .order_by(AiTcSuite.sort_order, AiTcSuite.id)
        )
        suites = list(rows.scalars().all())

        suite_case_pending: dict[int, set[int]] = {}
        if suites:
            all_suite_ids = [s.id for s in suites]
            pending_rows = await self.db.execute(
                select(AiTcCase.id, AiTcCase.suite_id)
                .join(AiTcTaskItem, AiTcTaskItem.case_id == AiTcCase.id)
                .where(
                    AiTcCase.suite_id.in_(all_suite_ids),
                    AiTcCase.is_deleted == 0,
                    AiTcTaskItem.confirm_status == ConfirmStatus.PENDING,
                    AiTcTaskItem.item_status == ItemStatus.SUCCESS,
                    AiTcTaskItem.is_deleted == 0,
                )
                .distinct()
            )
            for case_id, suite_id in pending_rows:
                if suite_id not in suite_case_pending:
                    suite_case_pending[suite_id] = set()
                suite_case_pending[suite_id].add(case_id)

        vo_map: dict[int, PendingSuiteNodeVO] = {}
        for s in suites:
            pending_cases = suite_case_pending.get(s.id, set())
            vo = PendingSuiteNodeVO(
                id=s.id, label=s.name, name=s.name,
                description=s.description,
                project_id=s.project_id, parent_id=s.parent_id,
                sort_order=s.sort_order,
                case_count=0,
                pending_count=len(pending_cases),
            )
            vo_map[s.id] = vo

        tree: list[PendingSuiteNodeVO] = []
        for vo in vo_map.values():
            parent = vo_map.get(vo.parent_id)
            if parent:
                parent.children.append(vo)
            else:
                tree.append(vo)

        def _accumulate(node: PendingSuiteNodeVO) -> int:
            for child in node.children:
                node.pending_count += _accumulate(child)
            return node.pending_count

        for root in tree:
            _accumulate(root)

        return tree

    async def get_pending_case_list(self, suite_id: int) -> list[PendingCaseVO]:
        """获取指定套件及其子树下所有待审核用例列表。"""
        rows = await self.db.execute(
            select(AiTcCase.id, AiTcCase.external_id, AiTcCase.name, AiTcCase.importance)
            .join(AiTcTaskItem, AiTcTaskItem.case_id == AiTcCase.id)
            .where(
                AiTcCase.suite_id == suite_id,
                AiTcCase.is_deleted == 0,
                AiTcTaskItem.confirm_status == ConfirmStatus.PENDING,
                AiTcTaskItem.item_status == ItemStatus.SUCCESS,
                AiTcTaskItem.is_deleted == 0,
            )
            .distinct()
            .order_by(AiTcCase.external_id, AiTcCase.id)
        )
        return [
            PendingCaseVO(id=r.id, external_id=r.external_id, name=r.name, importance=r.importance)
            for r in rows
        ]

    async def get_case_review_detail(self, case_id: int) -> CaseReviewDetailVO:
        """获取用例审核详情：原用例 + 最新 AI 建议。"""
        case = (await self.db.execute(
            select(AiTcCase).where(AiTcCase.id == case_id, AiTcCase.is_deleted == 0)
        )).scalar_one_or_none()

        suite_name = ""
        if case and case.suite_id:
            s = await self.db.get(AiTcSuite, case.suite_id)
            if s:
                suite_name = s.name

        project_prefix = ""
        if case and case.project_id:
            proj = await self.db.get(AiTcProject, case.project_id)
            if proj:
                project_prefix = proj.prefix

        case_vo = self._case_to_vo(case, suite_name, project_prefix) if case else None

        task_item_row = (await self.db.execute(
            select(AiTcTaskItem)
            .where(
                AiTcTaskItem.case_id == case_id,
                AiTcTaskItem.confirm_status == ConfirmStatus.PENDING,
                AiTcTaskItem.item_status == ItemStatus.SUCCESS,
                AiTcTaskItem.is_deleted == 0,
            )
            .order_by(AiTcTaskItem.id.desc())
            .limit(1)
        )).scalar_one_or_none()

        task_item_id = None
        task_id = None
        score: int | None = None
        issues: list[str] = []
        suggestions: list[FieldSuggestionVO] = []
        overall_assessment: str = ""

        if task_item_row and task_item_row.output:
            output: dict = task_item_row.output or {}
            task_item_id = task_item_row.id
            task_id = task_item_row.task_id
            score = output.get("score")
            issues = output.get("issues") or []
            overall_assessment = output.get("overall_assessment") or output.get("suggestion") or ""

            fields_data = output.get("fields") or []

            if fields_data and case:
                for fd in fields_data:
                    if not isinstance(fd, dict):
                        continue
                    fn = fd.get("field_name", "")
                    conclusion = fd.get("conclusion", "pass")
                    rule_violated = fd.get("rule_violated", "")
                    sv = fd.get("suggested_value")

                    original = self._get_case_field(case, fn)
                    has = conclusion == "fail" and sv is not None and str(sv) != str(original or "")

                    suggestions.append(FieldSuggestionVO(
                        field_name=fn,
                        original=original,
                        suggested=sv if has else None,
                        has_suggestion=has,
                        conclusion=conclusion,
                        rule_violated=rule_violated,
                    ))

            elif not fields_data and case:
                rewritten = output.get("rewritten") or {}
                if isinstance(rewritten, dict):
                    field_defs = [
                        ("name", "用例名称", case.name, rewritten.get("name")),
                        ("purpose", "测试目的", case.purpose, rewritten.get("purpose")),
                        ("summary", "测试思想", case.summary, rewritten.get("summary")),
                        ("preconditions", "前置条件", case.preconditions, rewritten.get("preconditions")),
                        ("test_data", "测试数据", case.test_data, rewritten.get("test_data")),
                        ("topo", "测试Topo", case.topo, rewritten.get("topo")),
                    ]
                    for field_name, _, original, suggested in field_defs:
                        has = suggested is not None and str(suggested) != str(original or "")
                        suggestions.append(FieldSuggestionVO(
                            field_name=field_name,
                            original=original,
                            suggested=suggested if has else None,
                            has_suggestion=has,
                            conclusion="fail" if has else "pass",
                            rule_violated="",
                        ))

                    ai_steps = rewritten.get("steps")
                    orig_steps = case.steps or []
                    if ai_steps is not None:
                        suggestions.append(FieldSuggestionVO(
                            field_name="steps",
                            original=orig_steps,
                            suggested=ai_steps,
                            has_suggestion=True,
                            conclusion="fail",
                            rule_violated="",
                        ))
                    else:
                        suggestions.append(FieldSuggestionVO(
                            field_name="steps",
                            original=orig_steps,
                            suggested=None,
                            has_suggestion=False,
                            conclusion="pass",
                            rule_violated="",
                        ))

        return CaseReviewDetailVO(
            case=case_vo, task_item_id=task_item_id, task_id=task_id,
            score=score, issues=issues, suggestions=suggestions,
            overall_assessment=overall_assessment,
        )

    async def review_case(
        self, req: CaseReviewReq, reviewed_by: str = "", reviewer_ip: str = ""
    ) -> None:
        """提交用例审核：逐字段采纳/忽略，更新用例，记录审计。"""
        item = await self.db.get(AiTcTaskItem, req.task_item_id)
        if item is None or item.case_id != req.case_id:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="任务明细不存在")

        case = await self.db.get(AiTcCase, req.case_id)
        if case is None:
            raise BusinessException(code=ResultCode.DATA_NOT_FOUND, msg="用例不存在")

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        output: dict = item.output or {}
        rewritten = output.get("rewritten") or {}
        has_accept = False

        for f in req.fields:
            if f.action == "accept" and f.field_name != "steps":
                suggested = rewritten.get(f.field_name)
                before_val = getattr(case, f.field_name, "") or ""
                if suggested is not None:
                    setattr(case, f.field_name, suggested)
                    has_accept = True

                self.db.add(AiTcReviewRecord(
                    task_id=item.task_id, task_item_id=item.id, case_id=req.case_id,
                    review_action="field_accept", field_name=f.field_name,
                    before_value=json.dumps(before_val, ensure_ascii=False),
                    after_value=json.dumps(suggested, ensure_ascii=False),
                    reviewer=reviewed_by, reviewer_ip=reviewer_ip, review_time=now,
                ))

            elif f.action == "accept" and f.field_name == "steps":
                ai_steps = rewritten.get("steps")
                before_val = json.dumps(case.steps or [], ensure_ascii=False)
                if ai_steps is not None:
                    case.steps = ai_steps
                    has_accept = True

                self.db.add(AiTcReviewRecord(
                    task_id=item.task_id, task_item_id=item.id, case_id=req.case_id,
                    review_action="field_accept", field_name="steps",
                    before_value=before_val,
                    after_value=json.dumps(ai_steps or [], ensure_ascii=False),
                    reviewer=reviewed_by, reviewer_ip=reviewer_ip, review_time=now,
                ))

            elif f.action == "edit_accept" and f.edited_value is not None:
                if f.field_name != "steps":
                    before_val = getattr(case, f.field_name, "") or ""
                    setattr(case, f.field_name, f.edited_value)
                    has_accept = True
                    self.db.add(AiTcReviewRecord(
                        task_id=item.task_id, task_item_id=item.id, case_id=req.case_id,
                        review_action="field_accept", field_name=f.field_name,
                        before_value=json.dumps(before_val, ensure_ascii=False),
                        after_value=json.dumps(f.edited_value, ensure_ascii=False),
                        reviewer=reviewed_by, reviewer_ip=reviewer_ip, review_time=now,
                        memo="manual_edit",
                    ))
                else:
                    before_val = json.dumps(case.steps or [], ensure_ascii=False)
                    case.steps = f.edited_value if isinstance(f.edited_value, list) else json.loads(f.edited_value)
                    has_accept = True
                    self.db.add(AiTcReviewRecord(
                        task_id=item.task_id, task_item_id=item.id, case_id=req.case_id,
                        review_action="field_accept", field_name="steps",
                        before_value=before_val,
                        after_value=json.dumps(f.edited_value, ensure_ascii=False),
                        reviewer=reviewed_by, reviewer_ip=reviewer_ip, review_time=now,
                        memo="manual_edit",
                    ))

            elif f.action == "ignore":
                before_val = getattr(case, f.field_name, "") if f.field_name != "steps" else json.dumps(case.steps or [], ensure_ascii=False)
                self.db.add(AiTcReviewRecord(
                    task_id=item.task_id, task_item_id=item.id, case_id=req.case_id,
                    review_action="ignore", field_name=f.field_name,
                    before_value=json.dumps(before_val, ensure_ascii=False) if not isinstance(before_val, str) else before_val,
                    after_value=None,
                    reviewer=reviewed_by, reviewer_ip=reviewer_ip, review_time=now,
                ))

        item.confirm_status = ConfirmStatus.ACCEPTED if has_accept else ConfirmStatus.IGNORED
        item.reviewed_by = reviewed_by
        item.review_time = now

        case.review_status = ReviewStatus.REVIEWED
        case.update_time = datetime.now()

        await self.db.flush()
        logger.info(f"Case {req.case_id} reviewed by {reviewed_by}, accepted fields: {sum(1 for f in req.fields if f.action == 'accept')}")

    # ═══════════════ Case 写操作（AI 结果写入） ═══════════════

    async def apply_core_select_result(self, case_id: int, reason: str = "", is_core: bool = True) -> None:
        """AI 核心用例挑选后写入标记。"""
        case = await self.db.get(AiTcCase, case_id)
        if case:
            case.is_core = 1 if is_core else 0
            case.core_reason = reason[:512] if reason else None
            case.core_source = CoreSource.AI
            case.update_time = datetime.now()

    async def apply_case_review_result(self, case_id: int, fields: dict) -> None:
        """将审核结果写入用例字段: {name, summary, preconditions, steps, test_data, topo}。"""
        case = await self.db.get(AiTcCase, case_id)
        if case is None:
            return
        if "name" in fields:
            case.name = fields["name"]
        if "purpose" in fields:
            case.purpose = fields["purpose"]
        if "summary" in fields:
            case.summary = fields["summary"]
        if "preconditions" in fields:
            case.preconditions = fields["preconditions"]
        if "steps" in fields:
            case.steps = fields["steps"]
        if "test_data" in fields:
            case.test_data = fields["test_data"]
        if "topo" in fields:
            case.topo = fields["topo"]
        case.review_status = 1
        case.update_time = datetime.now()

    async def mark_case_reviewed(self, case_id: int) -> None:
        case = await self.db.get(AiTcCase, case_id)
        if case:
            case.review_status = 1
            case.update_time = datetime.now()

    async def update_case_field(self, case_id: int, field_name: str, value) -> None:
        case = await self.db.get(AiTcCase, case_id)
        if case is None:
            return
        field_map = {
            "name": "name", "summary": "summary", "preconditions": "preconditions",
            "test_data": "test_data", "topo": "topo", "steps": "steps",
        }
        attr = field_map.get(field_name)
        if attr:
            setattr(case, attr, value)
            case.update_time = datetime.now()

    async def increment_case_script_count(self, case_id: int) -> None:
        case = await self.db.get(AiTcCase, case_id)
        if case:
            case.script_count = (case.script_count or 0) + 1
            case.update_time = datetime.now()

    # ═══════════════ VO 组装 ═══════════════

    def _project_to_vo(self, p: AiTcProject) -> ProjectVO:
        return ProjectVO(
            id=p.id, name=p.name, prefix=p.prefix, description=p.description,
            last_sync_time=p.last_sync_time,
            create_time=str(p.create_time) if p.create_time else None,
            update_time=str(p.update_time) if p.update_time else None,
        )

    @staticmethod
    def _get_case_field(case, field_name: str):
        """获取用例的字段值，用于构建审核详情。"""
        if field_name == "steps":
            return case.steps or []
        return getattr(case, field_name, "") or ""

    def _case_to_vo(self, c: AiTcCase, suite_name: str = "", project_prefix: str = "") -> CaseVO:
        steps = []
        if c.steps:
            steps = [CaseStep(**s) if isinstance(s, dict) else s for s in c.steps]
        return CaseVO(
            id=c.id, project_id=c.project_id, project_prefix=project_prefix,
            suite_id=c.suite_id, suite_name=suite_name,
            external_id=c.external_id, name=c.name, purpose=c.purpose,
            summary=c.summary, preconditions=c.preconditions,
            topo=c.topo, test_data=c.test_data,
            steps=steps,
            summary_raw=c.summary_raw, preconditions_raw=c.preconditions_raw,
            steps_raw=c.steps_raw, test_data_raw=c.test_data_raw,
            steps_parse_status=c.steps_parse_status,
            importance=c.importance,
            is_core=c.is_core, core_reason=c.core_reason, core_source=c.core_source,
            is_sample=c.is_sample,
            review_status=c.review_status, script_count=c.script_count,
            create_time=str(c.create_time) if c.create_time else None,
            update_time=str(c.update_time) if c.update_time else None,
        )
